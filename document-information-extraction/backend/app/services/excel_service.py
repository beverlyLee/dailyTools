import io
from typing import List, Dict, Any
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

class ExcelService:
    def __init__(self):
        self.header_font = Font(bold=True, size=12)
        self.header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        self.header_font_white = Font(bold=True, size=12, color="FFFFFF")
        self.thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    def export_invoices_to_excel(self, invoices: List[Any]) -> io.BytesIO:
        output = io.BytesIO()
        wb = Workbook()
        ws = wb.active
        ws.title = "增值税发票报销明细"
        
        headers = [
            "序号", "发票代码", "发票号码", "开票日期", "金额", "税额", "价税合计",
            "销售方名称", "销售方税号", "购买方名称", "购买方税号",
            "验真状态", "是否报销", "录入时间"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.header_font_white
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.thin_border
        
        for row_idx, invoice in enumerate(invoices, 2):
            row_data = [
                row_idx - 1,
                invoice.invoice_code or "",
                invoice.invoice_number or "",
                invoice.invoice_date.strftime("%Y-%m-%d") if invoice.invoice_date else "",
                invoice.amount or 0.0,
                invoice.tax_amount or 0.0,
                invoice.total_amount or 0.0,
                invoice.seller_name or "",
                invoice.seller_tax_id or "",
                invoice.buyer_name or "",
                invoice.buyer_tax_id or "",
                "已验真" if invoice.is_verified else "未验真",
                "已报销" if invoice.is_reimbursed else "未报销",
                invoice.created_at.strftime("%Y-%m-%d %H:%M:%S") if invoice.created_at else ""
            ]
            
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = self.thin_border
                
                if col_idx in [5, 6, 7] and isinstance(value, (int, float)):
                    cell.number_format = '¥#,##0.00'
        
        for col_idx in range(1, len(headers) + 1):
            column_letter = get_column_letter(col_idx)
            max_length = len(headers[col_idx - 1])
            for row_idx in range(1, len(invoices) + 2):
                cell = ws[f"{column_letter}{row_idx}"]
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[column_letter].width = min(max_length + 2, 30)
        
        wb.save(output)
        output.seek(0)
        return output
    
    def export_train_tickets_to_excel(self, tickets: List[Any]) -> io.BytesIO:
        output = io.BytesIO()
        wb = Workbook()
        ws = wb.active
        ws.title = "火车票报销明细"
        
        headers = [
            "序号", "车票号码", "出发站", "到达站", "出发时间", "车次",
            "座位类型", "票价", "乘客姓名", "身份证号",
            "验真状态", "是否报销", "录入时间"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.header_font_white
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.thin_border
        
        for row_idx, ticket in enumerate(tickets, 2):
            row_data = [
                row_idx - 1,
                ticket.ticket_number or "",
                ticket.departure_station or "",
                ticket.arrival_station or "",
                ticket.departure_time.strftime("%Y-%m-%d %H:%M") if ticket.departure_time else "",
                ticket.train_number or "",
                ticket.seat_class or "",
                ticket.price or 0.0,
                ticket.passenger_name or "",
                ticket.id_number or "",
                "已验真" if ticket.is_verified else "未验真",
                "已报销" if ticket.is_reimbursed else "未报销",
                ticket.created_at.strftime("%Y-%m-%d %H:%M:%S") if ticket.created_at else ""
            ]
            
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = self.thin_border
                
                if col_idx == 8 and isinstance(value, (int, float)):
                    cell.number_format = '¥#,##0.00'
        
        for col_idx in range(1, len(headers) + 1):
            column_letter = get_column_letter(col_idx)
            max_length = len(headers[col_idx - 1])
            for row_idx in range(1, len(tickets) + 2):
                cell = ws[f"{column_letter}{row_idx}"]
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[column_letter].width = min(max_length + 2, 30)
        
        wb.save(output)
        output.seek(0)
        return output
    
    def export_air_tickets_to_excel(self, tickets: List[Any]) -> io.BytesIO:
        output = io.BytesIO()
        wb = Workbook()
        ws = wb.active
        ws.title = "机票报销明细"
        
        headers = [
            "序号", "机票号码", "出发机场", "到达机场", "出发时间", "航班号",
            "舱位类型", "票价", "税费", "合计金额", "乘客姓名", "身份证号",
            "航空公司", "验真状态", "是否报销", "录入时间"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.header_font_white
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.thin_border
        
        for row_idx, ticket in enumerate(tickets, 2):
            row_data = [
                row_idx - 1,
                ticket.ticket_number or "",
                ticket.departure_airport or "",
                ticket.arrival_airport or "",
                ticket.departure_time.strftime("%Y-%m-%d %H:%M") if ticket.departure_time else "",
                ticket.flight_number or "",
                ticket.cabin_class or "",
                ticket.price or 0.0,
                ticket.tax or 0.0,
                ticket.total_amount or 0.0,
                ticket.passenger_name or "",
                ticket.id_number or "",
                ticket.airline or "",
                "已验真" if ticket.is_verified else "未验真",
                "已报销" if ticket.is_reimbursed else "未报销",
                ticket.created_at.strftime("%Y-%m-%d %H:%M:%S") if ticket.created_at else ""
            ]
            
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = self.thin_border
                
                if col_idx in [8, 9, 10] and isinstance(value, (int, float)):
                    cell.number_format = '¥#,##0.00'
        
        for col_idx in range(1, len(headers) + 1):
            column_letter = get_column_letter(col_idx)
            max_length = len(headers[col_idx - 1])
            for row_idx in range(1, len(tickets) + 2):
                cell = ws[f"{column_letter}{row_idx}"]
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[column_letter].width = min(max_length + 2, 30)
        
        wb.save(output)
        output.seek(0)
        return output

excel_service = ExcelService()
