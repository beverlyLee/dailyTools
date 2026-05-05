import io
import os
from datetime import datetime
from decimal import Decimal
from typing import List, Optional
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, and_

from database import Invoice, InvoiceItem, Company
from config import settings

class ExcelExporter:
    def __init__(self):
        self._workbook = None
    
    async def export_invoices(
        self,
        invoices: List[Invoice],
        include_items: bool = True,
        include_verification: bool = True
    ) -> bytes:
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
            from openpyxl.utils import get_column_letter
        except ImportError:
            return await self._export_with_xlsxwriter(invoices, include_items, include_verification)
        
        wb = openpyxl.Workbook()
        
        header_font = Font(bold=True, size=11)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font_white = Font(bold=True, size=11, color="FFFFFF")
        center_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        ws_summary = wb.active
        ws_summary.title = "发票汇总"
        
        summary_headers = [
            "序号", "票据类型", "发票代码", "发票号码", "开票日期",
            "销售方名称", "销售方税号", "购买方名称", "购买方税号",
            "金额", "税额", "价税合计", "核验状态", "报销状态", "备注"
        ]
        
        for col, header in enumerate(summary_headers, 1):
            cell = ws_summary.cell(row=1, column=col, value=header)
            cell.font = header_font_white
            cell.fill = header_fill
            cell.alignment = center_alignment
            cell.border = thin_border
        
        for row_idx, invoice in enumerate(invoices, 2):
            invoice_type_map = {
                "vat_invoice": "增值税发票",
                "train_ticket": "火车票",
                "flight_ticket": "机票",
                "receipt": "其他票据"
            }
            
            row_data = [
                row_idx - 1,
                invoice_type_map.get(invoice.invoice_type, invoice.invoice_type),
                invoice.invoice_code or "",
                invoice.invoice_number or "",
                invoice.invoice_date.strftime("%Y-%m-%d") if invoice.invoice_date else "",
                invoice.seller_name or "",
                invoice.seller_tax_id or "",
                invoice.buyer_name or "",
                invoice.buyer_tax_id or "",
                float(invoice.total_amount) if invoice.total_amount else 0,
                float(invoice.total_tax) if invoice.total_tax else 0,
                float(invoice.total_amount_with_tax) if invoice.total_amount_with_tax else 0,
                "已核验" if invoice.is_verified else "待核验",
                "已报销" if invoice.is_reimbursed else "未报销",
                invoice.remarks or ""
            ]
            
            for col, value in enumerate(row_data, 1):
                cell = ws_summary.cell(row=row_idx, column=col, value=value)
                cell.border = thin_border
                
                if col in [10, 11, 12]:
                    cell.number_format = '#,##0.00'
        
        for col_idx in range(1, len(summary_headers) + 1):
            ws_summary.column_dimensions[get_column_letter(col_idx)].width = 18
        
        if include_items:
            ws_items = wb.create_sheet("明细项目")
            
            item_headers = [
                "序号", "发票代码", "发票号码", "项目名称", "规格型号",
                "单位", "数量", "单价", "金额", "税率", "税额"
            ]
            
            for col, header in enumerate(item_headers, 1):
                cell = ws_items.cell(row=1, column=col, value=header)
                cell.font = header_font_white
                cell.fill = header_fill
                cell.alignment = center_alignment
                cell.border = thin_border
            
            item_row = 2
            item_count = 1
            
            for invoice in invoices:
                if hasattr(invoice, 'items') and invoice.items:
                    for item in invoice.items:
                        row_data = [
                            item_count,
                            invoice.invoice_code or "",
                            invoice.invoice_number or "",
                            item.item_name or "",
                            item.specification or "",
                            item.unit or "",
                            float(item.quantity) if item.quantity else 0,
                            float(item.unit_price) if item.unit_price else 0,
                            float(item.amount) if item.amount else 0,
                            item.tax_rate or "",
                            float(item.tax_amount) if item.tax_amount else 0
                        ]
                        
                        for col, value in enumerate(row_data, 1):
                            cell = ws_items.cell(row=item_row, column=col, value=value)
                            cell.border = thin_border
                            
                            if col in [7, 8, 9, 11]:
                                cell.number_format = '#,##0.00'
                        
                        item_row += 1
                        item_count += 1
            
            for col_idx in range(1, len(item_headers) + 1):
                ws_items.column_dimensions[get_column_letter(col_idx)].width = 15
        
        ws_stats = wb.create_sheet("统计")
        
        stats_data = [
            ["统计项", "值"],
            ["导出时间", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
            ["发票总数", len(invoices)],
            ["总金额", sum(float(inv.total_amount) if inv.total_amount else 0 for inv in invoices)],
            ["总税额", sum(float(inv.total_tax) if inv.total_tax else 0 for inv in invoices)],
            ["价税合计", sum(float(inv.total_amount_with_tax) if inv.total_amount_with_tax else 0 for inv in invoices)],
            ["已核验", sum(1 for inv in invoices if inv.is_verified)],
            ["已报销", sum(1 for inv in invoices if inv.is_reimbursed)]
        ]
        
        for row_idx, row_data in enumerate(stats_data, 1):
            for col, value in enumerate(row_data, 1):
                cell = ws_stats.cell(row=row_idx, column=col, value=value)
                if row_idx == 1:
                    cell.font = header_font_white
                    cell.fill = header_fill
                cell.border = thin_border
        
        ws_stats.column_dimensions['A'].width = 20
        ws_stats.column_dimensions['B'].width = 30
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output.read()
    
    async def _export_with_xlsxwriter(
        self,
        invoices: List[Invoice],
        include_items: bool = True,
        include_verification: bool = True
    ) -> bytes:
        import xlsxwriter
        
        output = io.BytesIO()
        workbook = xlsxwriter.Workbook(output)
        
        header_format = workbook.add_format({
            'bold': True,
            'bg_color': '#4472C4',
            'font_color': 'white',
            'align': 'center',
            'valign': 'vcenter',
            'border': 1
        })
        
        cell_format = workbook.add_format({
            'align': 'center',
            'valign': 'vcenter',
            'border': 1
        })
        
        money_format = workbook.add_format({
            'num_format': '#,##0.00',
            'align': 'center',
            'valign': 'vcenter',
            'border': 1
        })
        
        ws_summary = workbook.add_worksheet("发票汇总")
        
        summary_headers = [
            "序号", "票据类型", "发票代码", "发票号码", "开票日期",
            "销售方名称", "销售方税号", "购买方名称", "购买方税号",
            "金额", "税额", "价税合计", "核验状态", "报销状态", "备注"
        ]
        
        for col, header in enumerate(summary_headers):
            ws_summary.write(0, col, header, header_format)
            ws_summary.set_column(col, col, 18)
        
        invoice_type_map = {
            "vat_invoice": "增值税发票",
            "train_ticket": "火车票",
            "flight_ticket": "机票",
            "receipt": "其他票据"
        }
        
        for row_idx, invoice in enumerate(invoices, 1):
            row_data = [
                row_idx,
                invoice_type_map.get(invoice.invoice_type, invoice.invoice_type),
                invoice.invoice_code or "",
                invoice.invoice_number or "",
                invoice.invoice_date.strftime("%Y-%m-%d") if invoice.invoice_date else "",
                invoice.seller_name or "",
                invoice.seller_tax_id or "",
                invoice.buyer_name or "",
                invoice.buyer_tax_id or "",
                float(invoice.total_amount) if invoice.total_amount else 0,
                float(invoice.total_tax) if invoice.total_tax else 0,
                float(invoice.total_amount_with_tax) if invoice.total_amount_with_tax else 0,
                "已核验" if invoice.is_verified else "待核验",
                "已报销" if invoice.is_reimbursed else "未报销",
                invoice.remarks or ""
            ]
            
            for col, value in enumerate(row_data):
                if col in [9, 10, 11]:
                    ws_summary.write(row_idx, col, value, money_format)
                else:
                    ws_summary.write(row_idx, col, value, cell_format)
        
        if include_items:
            ws_items = workbook.add_worksheet("明细项目")
            
            item_headers = [
                "序号", "发票代码", "发票号码", "项目名称", "规格型号",
                "单位", "数量", "单价", "金额", "税率", "税额"
            ]
            
            for col, header in enumerate(item_headers):
                ws_items.write(0, col, header, header_format)
                ws_items.set_column(col, col, 15)
            
            item_row = 1
            item_count = 1
            
            for invoice in invoices:
                if hasattr(invoice, 'items') and invoice.items:
                    for item in invoice.items:
                        row_data = [
                            item_count,
                            invoice.invoice_code or "",
                            invoice.invoice_number or "",
                            item.item_name or "",
                            item.specification or "",
                            item.unit or "",
                            float(item.quantity) if item.quantity else 0,
                            float(item.unit_price) if item.unit_price else 0,
                            float(item.amount) if item.amount else 0,
                            item.tax_rate or "",
                            float(item.tax_amount) if item.tax_amount else 0
                        ]
                        
                        for col, value in enumerate(row_data):
                            if col in [6, 7, 8, 10]:
                                ws_items.write(item_row, col, value, money_format)
                            else:
                                ws_items.write(item_row, col, value, cell_format)
                        
                        item_row += 1
                        item_count += 1
        
        workbook.close()
        output.seek(0)
        
        return output.read()

async def export_invoices_to_excel(
    invoices: List[Invoice],
    include_items: bool = True,
    include_verification: bool = True
) -> bytes:
    exporter = ExcelExporter()
    return await exporter.export_invoices(invoices, include_items, include_verification)

def generate_batch_number() -> str:
    now = datetime.now()
    return f"RB{now.strftime('%Y%m%d%H%M%S')}"
